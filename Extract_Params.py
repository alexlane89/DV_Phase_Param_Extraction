import subprocess
import os
import tkinter
from openpyxl import Workbook

# select file from dialog
tkinter.Tk().withdraw() # prevent tk "root window" from appearing with file dialog
from tkinter.filedialog import askopenfilename
filepath = askopenfilename()

# generate_params function to search file path for parameter type
# and generate resulting list of line entries which
# include the specific parameter type.
def generate_params(param_type):
    with open(filepath, encoding="utf16") as fp:
        linelist = fp.readlines()
        result_set = []
        for line in linelist:
            if param_type in line:
                result_set.append(line)
    return result_set

# Generate list of parameters by splitting line
# entries resulting from generate_params function
# and returning parameter name.
type_list = ['BATCH_REPORT_INTEGER', 'INT8', 'INT16', 'INT32']
new_list = []
for a in type_list:
    for b in generate_params(a):
        new_list.append(b.split('"')[1:2])
    flat_list = []
    for sublist in new_list: #new_list is generated as a list of lists. Convert to flat_list to simply list items.
        for item in sublist:
            flat_list.append(item)

# Generate list of lines which contain the parameter(s) in question
with open(filepath, encoding="utf16") as fp:
    linelist = fp.readlines()
    result_actions = []
    for line in linelist:
        for c in flat_list:
            if c in line:
                result_actions.append(line)

# write expressions referencing parameters in question
# to a .xlsx file
file_name = os.path.split(filepath)[1].split(".")[0] + '.xlsx'
wb = Workbook()
ws = wb.active
ws.cell(row=1,column = 1).value = 'Expressions'
for j in range(len(result_actions)):
    ws.cell(row=j+2,column = 1).value = result_actions[j]
wb.save(file_name)